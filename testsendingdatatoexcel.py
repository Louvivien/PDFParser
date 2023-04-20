import openpyxl
import re
import ast

def extract_postal_code(address):
    match = re.search(r'\b\d{5}\b', address)
    if match:
        return match.group(0)[:2]
    return "99"

def add_row(sheet, data):
    row = sheet.max_row + 1
    for col, value in data.items():
        sheet.cell(row=row, column=col).value = value
    return row

def process_data(file_path, data_list):
    wb = openpyxl.load_workbook(file_path)

    for data in data_list:
        if data["Type"] == "RECEPTION":
            sheet = wb["GRAINES"]
            row_data = {
                1: data["Date entree"],
                2: data["Fournisseur"],
                7: data["Ticket number"],
                8: data["Immatriculation camion"],
                10: data["Numero contrat"],
                12: data["Poids net"],
            }
            row = add_row(sheet, row_data)
            print(f"Updated row {row} in GRAINES tab of {file_path}")

        elif data["Type"] == "EXPEDITION":
            if data["Produit"] == "HUILE":
                sheet = wb["HUILE"]
                postal_code = extract_postal_code(data["Adresse client"])
                client = f"{postal_code}{data['Client']}"
                row_data = {
                    1: data["Date de sortie"],
                    6: client,
                    11: data["Ticket number"],
                    12: data["Immatriculation camion"],
                    13: data["Numero contrat"],
                    15: data["Poids net"],
                }
                row = add_row(sheet, row_data)
                print(f"Updated row {row} in HUILE tab of {file_path}")

            elif data["Produit"] == "TOURTEAU":
                sheet = wb["TOURTEAUX"]
                postal_code = extract_postal_code(data["Adresse client"])
                client = f"{postal_code}{data['Client']}"
                if data["Client"] == "AXEREAL FE":
                    client = "AXEREAL/03THIVAT"
                row_data = {
                    1: data["Date de sortie"],
                    2: client,
                    5: data["Ticket number"],
                    6: data["Immatriculation camion"],
                    7: data["Numero contrat"],
                    9: data["Poids net"],
                }
                row = add_row(sheet, row_data)
                print(f"Updated row {row} in TOURTEAUX tab of {file_path}")

    wb.save(file_path)

def parse_input_data(input_data):
    input_data = input_data.strip().replace("\n", "").replace("'", "\"")
    return ast.literal_eval(input_data)

input_data = input("Enter the data: ")
data_dict = parse_input_data(input_data)
data_list = [data_dict]

file_path = "./data/sample.xlsx"
process_data(file_path, data_list)
